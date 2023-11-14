import cx_Oracle
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment

host = 'x'
servico = 'x'
usuario = 'x'
senha = 'x'

# Encontra o arquivo que aponta para o banco de dados
cx_Oracle.init_oracle_client(lib_dir="./instantclient_21_10")

# Faz a conexão ao banco de dados
conecta_banco = cx_Oracle.connect(usuario, senha, f'{host}/{servico}')

# Cria um cursor no banco para que seja possível fazer consultas e alterações no banco de dados
cursor = conecta_banco.cursor()

print("""   __________  __  _______  ____  ___   _____
  / ____/ __ \/  |/  / __ \/ __ \/   | / ___/
 / /   / / / / /|_/ / /_/ / /_/ / /| | \__ \ 
/ /___/ /_/ / /  / / ____/ _, _/ ___ |___/ / 
\____/\____/_/  /_/_/   /_/ |_/_/  |_/____/  
                                             """)

print('Informe as filiais separadas por vírgula')
filiais = input('Digite aqui:')
#print(filiais)

print('Tipo do pedido, BONIFICADO, VENDA ou TODOS')
print('Valores válidos: B, V ou T')
tipo_pedido = str(input('Digite aqui: '))
if tipo_pedido == 'B':
    tipo_pedido = "'BONIFICADO'"
elif tipo_pedido == 'V':
    tipo_pedido = "'VENDA'"
elif tipo_pedido == 'T':
    tipo_pedido = 'BONIFICADO','VENDA'
    #print(tipo_pedido)

print('Informe o status do pedido, entrega TOTAL, PARCIAL, PENDENTE, ou TODOS')
print('Valores validos: TOT, PAR, PEN, TOD')
status_entrega = str(input('Digite aqui: '))
if status_entrega == 'TOT':
    status_entrega = "'TOTAL'"
elif status_entrega == 'PAR':
    status_entrega = "'PARCIAL'"
elif status_entrega == 'PEN':
    status_entrega = "'PENDENTE'"
elif status_entrega == 'TOD':
    status_entrega = 'TOTAL','PARCIAL','PENDENTE'
    #print(status_entrega)

print('Informe a data inicial')
print('A data precisa estar neste formato: DD-MES-AAAA ex: 01-oct-2023')
print('IMPORTANTE: AS INICIAIS DO MÊS PRECISAM ESTAR EM INGLÊS')
data_inicial = str(input('Digite aqui: '))

print('Informe a data final')
data_final = str(input('Digite aqui: '))

print('Informe o código do comprador')
comprador = int(input('Digite aqui: '))



cursor.execute("""SELECT
    codfornec,
    fornecedor,
    numped,
    dtemissao,
    vltotal,
    codfilial,
    codcomprador,
    LISTAGG(NUMNOTA, ', ') WITHIN GROUP (ORDER BY NUMNOTA) AS NOTAS_FISCAIS,
    MAX(status_entrega) AS status_entrega,
    MAX(tipo_pedido) AS tipo_pedido
FROM (
    SELECT
        p.codfornec,
        f.fornecedor,
        p.numped,
        p.dtemissao,
        p.vltotal,
        p.codfilial,
        p.codcomprador,
        N.NUMNOTA,
        CASE
            WHEN p.vltotal = p.vlentregue THEN 'TOTAL'
            WHEN p.vlentregue = 0 THEN 'PENDENTE'
            ELSE 'PARCIAL'
        END AS status_entrega,
        CASE
            WHEN p.tipobonific = 'B' THEN 'BONIFICADO'
            WHEN p.tipobonific = 'N' THEN 'VENDA'
        END AS tipo_pedido
    FROM
        pcpedido p
        LEFT JOIN pcfornec f ON p.codfornec = f.codfornec
        LEFT JOIN PCPEDNF PN ON p.numped = PN.numpedido
        LEFT JOIN PCNFENT N ON PN.numtransent = N.numtransent
    WHERE
        p.codfilial IN ({})
        AND p.dtemissao BETWEEN TO_DATE('{}', 'DD-MON-YYYY') AND TO_DATE('{}', 'DD-MON-YYYY')
        AND p.codcomprador = {}
        --AND N.ESPECIE = 'NF'
) 
WHERE
    status_entrega IN {}
    AND tipo_pedido IN {}
GROUP BY
    codfornec, fornecedor, numped, dtemissao, vltotal, codfilial, codcomprador
ORDER BY
    dtemissao
                """.format(filiais, data_inicial, data_final, comprador, status_entrega, tipo_pedido,))

resultado = cursor.fetchall()

# Obter as descrições das colunas da consulta
column_descriptions = [desc[0] for desc in cursor.description]

    # Filtrar as colunas que não contêm objetos LOB
filtered_results = []
for row in resultado:
    filtered_row = [str(cell) if not isinstance(cell, cx_Oracle.LOB) else '' for cell in row]
    filtered_results.append(filtered_row)

    # Converter os resultados em um DataFrame do Pandas
df = pd.DataFrame(filtered_results, columns=column_descriptions)

    # Especifique o nome do arquivo Excel de saída
output_file = "resultado_da_consulta.xlsx"

# Salvar o DataFrame no arquivo Excel, excluindo a coluna "COMPRADOR"
df.drop(columns=['CODCOMPRADOR'], inplace=True)
df.to_excel(output_file, index=False, engine='openpyxl')

# Abra o arquivo Excel usando openpyxl
wb = openpyxl.load_workbook(output_file)
ws = wb.active



# Defina um padrão de preenchimento verde
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Percorra as células da coluna "status_entrega" e aplique a formatação verde se o valor for "TOTAL"
for row in range(2, ws.max_row + 1):
    cell_value = ws.cell(row=row, column=8).value  # Coluna 8 corresponde à "status_entrega"
    if cell_value == 'TOTAL':
        ws.cell(row=row, column=8).fill = green_fill


# Defina um padrão de preenchimento azul
blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

# Percorra as células da coluna "status_entrega" e aplique a formatação azul se o valor for "PARCIAL"
for row in range(2, ws.max_row + 1):
    cell_value = ws.cell(row=row, column=8).value  # Coluna 8 corresponde à "status_entrega"
    if cell_value == 'PARCIAL':
        ws.cell(row=row, column=8).fill = blue_fill


# Defina um padrão de preenchimento vermelho
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Percorra as células da coluna "status_entrega" e aplique a formatação vermelho se o valor for "PENDENTE"
for row in range(2, ws.max_row + 1):
    cell_value = ws.cell(row=row, column=8).value  # Coluna 8 corresponde à "status_entrega"
    if cell_value == 'PENDENTE':
        ws.cell(row=row, column=8).fill = red_fill
        

# Substituir o ponto por vírgula na coluna "vltotal"
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=5)  # Coluna 5 corresponde à "vltotal"
    # Obtém o valor da célula como texto e faz a substituição do ponto por vírgula
    cell.value = str(cell.value).replace('.', ',')
    # Tenta converter o valor para float, se não for possível, mantém o valor original
    try:
        cell.value = float(cell_value)
        # Defina o formato numérico
        cell.number_format = '0.00'
    except ValueError:
        pass  # Mantém o valor original se não for um número

# Formatar a coluna "dtemissão" no arquivo Excel para mostrar apenas a data
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=4)  # Coluna 4 corresponde à "dtemissão"
    cell.number_format = 'DD/MM/YYYY'

# Centralize todas as células na planilha
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
# Ajustar o tamanho das colunas automaticamente com base no conteúdo
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Salve as alterações no arquivo Excel
wb.save(output_file)

print(f"Consulta concluída com sucesso. Resultados salvos em {output_file}")




