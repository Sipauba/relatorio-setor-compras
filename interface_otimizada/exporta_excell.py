import cx_Oracle
from variaveis import dados
from conecta_banco import cursor
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment

def x():
    from variaveis import sql_geral
    print(sql_geral)

def exporta_excell():
    
  
    # Obter as descrições das colunas da consulta
    column_descriptions = [desc[0] for desc in cursor.description]

        # Filtrar as colunas que não contêm objetos LOB
    filtered_results = []
    for row in dados:
        filtered_row = [str(cell) if not isinstance(cell, cx_Oracle.LOB) else '' for cell in row]
        filtered_results.append(filtered_row)

        # Converter os resultados em um DataFrame do Pandas
    df = pd.DataFrame(filtered_results, columns=column_descriptions)

        # Especifique o nome do arquivo Excel de saída
    output_file = "resultado_da_consulta.xlsx"

    """# Salvar o DataFrame no arquivo Excel, excluindo a coluna "COMPRADOR"
    df.drop(columns=['CODCOMPRADOR'], inplace=True)
    df.to_excel(output_file, index=False, engine='openpyxl')"""

    # Abra o arquivo Excel usando openpyxl
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active



    # Defina um padrão de preenchimento verde
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Percorra as células da coluna "status_entrega" e aplique a formatação verde se o valor for "TOTAL"
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=8).value  # Coluna 8 corresponde à "status_entrega"
        if cell_value == 'TOTAL':
            ws.cell(row=row, column=8).fill = green_fill


    # Defina um padrão de preenchimento azul
    blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    # Percorra as células da coluna "status_entrega" e aplique a formatação azul se o valor for "PARCIAL"
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=8).value  # Coluna 8 corresponde à "status_entrega"
        if cell_value == 'PARCIAL':
            ws.cell(row=row, column=8).fill = blue_fill


    # Defina um padrão de preenchimento vermelho
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Percorra as células da coluna "status_entrega" e aplique a formatação vermelho se o valor for "AGUARDANDO FATURAMENTO"
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=8).value  # Coluna 8 corresponde à "status_entrega"
        if cell_value == 'AGUARDANDO FATURAMENTO':
            ws.cell(row=row, column=8).fill = red_fill
            
            
    # Defina um padrão de preenchimento amarelo
    yellow_fill = PatternFill(start_color="FFFF65", end_color="FFFF65", fill_type="solid")

    # Percorra as células da coluna "status_entrega" e aplique a formatação amarelo se o valor for "AGUARDANDO ENTREGA"
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=8).value  # Coluna 8 corresponde à "status_entrega"
        if cell_value == 'AGUARDANDO ENTREGA':
            ws.cell(row=row, column=8).fill = yellow_fill
            

    # Substituir o ponto por vírgula na coluna "vltotal"
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=6)  # Coluna 6 corresponde à "vltotal"
        # Obtém o valor da célula como texto e faz a substituição do ponto por vírgula
        cell.value = str(cell.value).replace('.', ',')
        # Tenta converter o valor para float, se não for possível, mantém o valor original
        try:
            cell.value = float(cell_value)
            # Defina o formato numérico
            cell.number_format = '0.00'
        except ValueError:
            pass  # Mantém o valor original se não for um número

    # Formatar a coluna "dtemissão" no arquivo Excel para mostrar apenas a data
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=0)  # Coluna 0 corresponde à "dtemissão"
        cell.number_format = 'DD/MM/YYYY'

    # Centralize todas as células na planilha
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # Ajustar o tamanho das colunas automaticamente com base no conteúdo
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
    # Salve as alterações no arquivo Excel
    wb.save(output_file)

    print(f"Consulta concluída com sucesso. Resultados salvos em {output_file}")